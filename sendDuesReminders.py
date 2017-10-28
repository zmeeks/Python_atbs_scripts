#! /usr/bin/env python3
#  sendDuesReminders.py -- Sends emails based on payment status in spreadsheet.
#  Usage:	./sendDuesReminders <password to gmail acct> 

"""
README:
Automate The Boring Stuff Python (Ch. 16 -- pg 376) Script
solved/copied by Z Meeks 10/04/17

Note:	project found on pg. 376 -- code given wasn't quite correct 
		the format given for body (line 38 below) as was had a few problems.
		Current code (as of today) works great.
Note:	replace sending_email value with your gmaail username
"""


import openpyxl, smtplib, sys

sending_gmail = '_________@gmail.com' # replace with your_gmail_acct@gmail.com
gmail_pword = sys.argv[1]

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row = 1, column = lastCol).value

# Check each members payment status
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
	payment = sheet.cell(row = r, column = lastCol).value
	if payment != 'paid':
		name = sheet.cell(row = r, column = 1).value
		email = sheet.cell(row = r, column = 2).value
		unpaidMembers[name] = email

# Log in to email account
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(sending_gmail, gmail_pword)

# Send out reminder emails
for name, email in unpaidMembers.items():
	subject = "Subject: %s dues unpaid" % (latestMonth)
	body = 	("To: %s \n"
			"Subject: %s dues unpaid. \n\n" 
			"Dear %s, \n"
			"Records show that you have not paid dues for %s.  Please make this payment as soon as possible. Thank you! "
		   	) % (email, latestMonth, name, latestMonth)	
	print('Sending email to %s...' % email)
	sendmailStatus = smtpObj.sendmail(sending_gmail, email, body)

	if sendmailStatus != {}:
		print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
